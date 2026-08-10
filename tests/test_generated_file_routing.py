"""Offline routing/evidence boundaries for user-ordered generated files."""

from __future__ import annotations

from typing import Any

import pytest

from friday.agent_runtime import (
    _ATTACHMENT_MAP_PREFIX,
    _FILE_CONVERSATION_GROUNDS_PREFIX,
    _FILE_WEB_SOURCE_LEDGER_PREFIX,
    AgentContext,
    AgentRuntime,
    _blocks_from_text,
    _file_context_ground_records,
    _file_kind_from_request,
    _is_direct_file_request,
    asks_for_the_web,
)
from friday.permissions import ActorContext


@pytest.mark.parametrize(
    ("prompt", "expected"),
    [
        ("Сделай таблицу из Word-файла.", "xlsx"),
        ("Оформи это таблицей.", "xlsx"),
        ("Выгрузи таблицу.", "xlsx"),
        ("По данным из документа Word сделай таблицу.", "xlsx"),
        ("Сделай Word-таблицу из Excel.", "docx"),
        ("Оформи таблицу в Word по данным из Excel.", "docx"),
        ("Создай PDF на основе таблицы Excel.", "pdf"),
        ("Сделай отчёт из таблицы Excel.", "docx"),
        ("Оформи данные из PDF в Excel.", "xlsx"),
        ("Конвертируй Word в PDF.", "pdf"),
        ("Экспортируй это в Excel.", "xlsx"),
        ("Сделай картинку по этим данным.", "png"),
        ("Сделай обычный отчёт.", "docx"),
        ("Как договаривались, сделай таблицу.", "xlsx"),
        ("Когда сможешь, оформи это в PDF.", "pdf"),
        ("Я уже сделал расчёты, оформи их таблицей.", "xlsx"),
    ],
)
def test_output_format_is_resolved_from_the_target_not_the_source(
    prompt: str,
    expected: str,
) -> None:
    assert _is_direct_file_request(prompt) is True
    assert _file_kind_from_request(prompt) == expected


@pytest.mark.parametrize(
    ("prompt", "expected"),
    [
        ("Сделай таблицу из файла report.docx.", "xlsx"),
        ("Сделай таблицу по Word-файлу.", "xlsx"),
        ("По этому Word-документу сделай таблицу.", "xlsx"),
        ("На базе Word-документа сделай таблицу.", "xlsx"),
        ("Используй Word-документ и сделай таблицу.", "xlsx"),
        ("Сделай отчёт по PDF.", "docx"),
        ("Сделай отчёт, исходник — Excel.", "docx"),
        ("Сделай не Word, а PDF.", "pdf"),
        ("Сделай не PDF, а Word.", "docx"),
        ("Конвертируй не в PDF, а в Word.", "docx"),
        ("Сделай таблицу из данных, сохрани в Word.", "docx"),
        ("Сделай отчёт на основе этих данных, оформи в PDF.", "pdf"),
        ("Сделай таблицу из данных, итог оформи в Word.", "docx"),
        ("Сделай таблицу с этими данными и оформи её в Word.", "docx"),
        ("Сделай отчёт на основе этих данных и оформи в PDF.", "pdf"),
        ("Сделай таблицу из данных и сохрани в Word.", "docx"),
    ],
)
def test_source_formats_and_rejected_targets_never_choose_the_output(prompt: str, expected: str) -> None:
    assert _file_kind_from_request(prompt) == expected


@pytest.mark.parametrize(
    "prompt",
    [
        "Как оформить таблицу?",
        "Где лучше оформить таблицу?",
        "Умеешь ли ты создавать Word-документы?",
        "Можно ли экспортировать таблицы в Excel?",
        "Я уже сделал таблицу в Excel.",
        "Ответь таблицей прямо в чате.",
        "Что находится в этой Excel-таблице?",
        "Переведи таблицу на английский язык.",
        "Сделай краткую сводку по таблице.",
    ],
)
def test_questions_narration_and_chat_formatting_do_not_authorize_a_file(
    prompt: str,
) -> None:
    assert _is_direct_file_request(prompt) is False


@pytest.mark.parametrize(
    "prompt",
    [
        "Ты умеешь создавать PDF?",
        "Поддерживаешь экспорт в PDF?",
        "Есть ли возможность экспортировать в PDF?",
        "Расскажи про экспорт в Excel.",
        "Объясни конвертацию Word в PDF.",
        "Что такое экспорт в Excel?",
        "Он попросил создать PDF.",
        "Покажи пример команды, которая создаёт PDF.",
        "Начальник попросил создать PDF.",
        "В инструкции сказано создать PDF.",
        "Мне сказали создать PDF.",
        "Подскажи, как создать PDF.",
        "Я не понимаю, как создать PDF.",
        "Помоги понять, как экспортировать таблицу в Excel.",
        "Хочу узнать, как конвертировать Word в PDF.",
        "Напиши инструкцию, как создать PDF.",
        "Есть вопрос: как создать PDF?",
        "Сделай краткую сводку по PDF.",
        "Сделай анализ Word-документа.",
        "Сделай выводы по Excel-файлу.",
        "Сделай краткий пересказ PDF.",
        "Составь список ошибок в PDF-файле.",
        "Подготовь ответ по Word-документу.",
    ],
)
def test_capability_explanations_and_reported_commands_do_not_create_files(prompt: str) -> None:
    assert _is_direct_file_request(prompt) is False


@pytest.mark.parametrize(
    "prompt",
    [
        "Я уже сделал расчёты. Теперь оформи их таблицей.",
        "Я уже создал Word. Конвертируй его в PDF.",
        "Дай таблицу.",
        "Нужна таблица.",
        "Я не хочу таблицу в чате, сделай Excel-файл.",
        "Покажи данные и создай PDF.",
        "Объясни расчёт и оформи его в PDF.",
        "Можешь сделать PDF?",
        "Расскажи, как устроен экспорт, и создай PDF.",
        "Объясни, как посчитано, и оформи в PDF.",
        "Что нужно учесть, чтобы всё работало? Затем создай PDF.",
        "Расскажи как устроен экспорт и создай PDF.",
        "Объясни как посчитано и оформи в PDF.",
        "Что нужно учесть чтобы всё работало и затем создай PDF.",
        "Я уже сделал расчёты и теперь оформи их таблицей.",
        "Сделай список в PDF.",
        "Составь список в Word.",
        "Подготовь список в Excel.",
        "Можешь оформить сводку в PDF?",
        "Сделай сводку в файле Word.",
        "Сделай список сотрудников в Excel.",
        "Сделай список людей в Excel.",
        "Составь список имён в Word.",
        "Подготовь список данных в PDF.",
        "Сделай список пунктов в PDF.",
        "Составь список фактов в Word.",
    ],
)
def test_a_later_or_direct_creation_command_is_authoritative(prompt: str) -> None:
    assert _is_direct_file_request(prompt) is True


@pytest.mark.parametrize(
    "prompt",
    [
        "Расскажи как создать PDF и отправь объяснение сообщением.",
        "Объясни как экспортировать в Excel и сохрани инструкцию в памяти.",
        "Подскажи как сделать PDF и пришли шаги текстом.",
        "Расскажи как создать PDF и отправь ответ в чат.",
    ],
)
def test_a_later_non_file_action_does_not_repurpose_an_earlier_format(prompt: str) -> None:
    assert _is_direct_file_request(prompt) is False


@pytest.mark.parametrize(
    "prompt",
    [
        "Сделай Excel по данным из интернета о ключевой ставке.",
        "Собери красивый PDF на основе информации из сети о проекте.",
        "Оформи Word из найденного в интернете материала.",
        "Выгрузи таблицу по результатам веб-поиска.",
    ],
)
def test_a_file_explicitly_based_on_public_web_data_authorizes_research(
    prompt: str,
) -> None:
    assert _is_direct_file_request(prompt) is True
    assert asks_for_the_web(prompt) is True


@pytest.mark.parametrize(
    "prompt",
    [
        "Сделай Excel с актуальными ценами из интернета.",
        "Сделай PDF про проект, информацию найди в интернете.",
        "Сделай Excel, найдя цены в интернете.",
        "Сделай Excel из информации, которую найдёшь в интернете.",
        "Сделай таблицу на основе страницы https://safe.synthetic.example.com/data.",
        "Собери PDF по этой веб-странице https://safe.synthetic.example.com/data.",
        "Сделай Excel по данным с сайта Росстата.",
        "Собери PDF на основе информации на сайте ЦБ.",
        "Выгрузи таблицу по данным портала Госуслуг.",
    ],
)
def test_natural_web_sourced_file_requests_authorize_research(prompt: str) -> None:
    assert asks_for_the_web(prompt) is True


def test_a_file_about_the_internet_is_not_automatically_a_web_request() -> None:
    prompt = "Сделай Word-памятку о том, как устроен интернет."
    assert _is_direct_file_request(prompt) is True
    assert asks_for_the_web(prompt) is False


@pytest.mark.parametrize(
    "prompt",
    [
        "Сделай PDF-визитку: укажи сайт https://company.example.org.",
        "Создай Word с одной строкой: https://company.example.org.",
        "Сделай таблицу, в колонку Сайт запиши https://company.example.org.",
    ],
)
def test_a_url_printed_as_literal_file_data_does_not_authorize_research(prompt: str) -> None:
    assert _is_direct_file_request(prompt) is True
    assert asks_for_the_web(prompt) is False


@pytest.mark.parametrize(
    "prompt",
    [
        "Сделай таблицу на основе страницы https://safe.synthetic.example.com/data.",
        "Собери PDF по данным веб-страницы https://safe.synthetic.example.com/data.",
        "Создай Word, изучив https://safe.synthetic.example.com/data.",
    ],
)
def test_a_url_named_as_the_file_source_authorizes_research(prompt: str) -> None:
    assert _is_direct_file_request(prompt) is True
    assert asks_for_the_web(prompt) is True


@pytest.mark.parametrize(
    ("prompt", "expected_query"),
    [
        ("Сделай Excel по данным из интернета о ключевой ставке.", "ключевой ставке"),
        ("Собери красивый PDF на основе информации из сети о проекте.", "проекте"),
        ("Оформи Word из найденного в интернете материала о ставке ЦБ.", "ставке ЦБ"),
        ("Выгрузи таблицу по результатам веб-поиска о ценах.", "ценах"),
    ],
)
def test_web_file_research_query_excludes_carrier_instructions(
    prompt: str,
    expected_query: str,
) -> None:
    assert AgentRuntime.web_query_from(prompt) == expected_query


@pytest.mark.parametrize(
    ("prompt", "expected_query"),
    [
        ("Найди в интернете цены и сделай Excel.", "цены"),
        ("Найди цены в интернете, затем сделай Excel.", "цены"),
        ("Сделай Excel с актуальными ценами из интернета.", "актуальными ценами"),
    ],
)
def test_combined_web_and_file_requests_keep_only_the_research_subject(
    prompt: str,
    expected_query: str,
) -> None:
    assert AgentRuntime.web_query_from(prompt) == expected_query


def test_a_markdown_table_becomes_real_excel_cells_without_losing_row_288() -> None:
    import io

    import openpyxl

    from friday.reports import render, spec_from_payload

    rows = ["| Позиция | Человек | Примечание |", "| ---: | --- | --- |"]
    rows.extend(f"| {index} | Человек {index} | Строка {index} |" for index in range(1, 301))
    blocks = _blocks_from_text("Реестр\n" + "\n".join(rows))

    assert len(blocks) == 1
    assert blocks[0]["kind"] == "table"
    assert len(blocks[0]["rows"]) == 301
    assert blocks[0]["rows"][288] == ["288", "Человек 288", "Строка 288"]

    workbook = openpyxl.load_workbook(io.BytesIO(render("xlsx", spec_from_payload("Реестр", "", blocks))))
    sheet = workbook.active
    header_row = next(cell.row for cell in sheet["A"] if cell.value == "Позиция")
    assert sheet.cell(header_row + 288, 1).value == "288"
    assert sheet.cell(header_row + 288, 2).value == "Человек 288"
    assert sheet.auto_filter.ref.endswith(f"{header_row + 300}")


def test_ragged_markdown_and_tabular_rows_remain_structured_tables() -> None:
    markdown = _blocks_from_text("| A | B |\n| --- | --- |\n| x |")
    assert markdown == [{"kind": "table", "rows": [["A", "B"], ["x", ""]]}]

    tab_rows = ["Позиция\tЧеловек\tПримечание"]
    tab_rows.extend(f"{index}\tЧеловек {index}\tСтрока {index}" for index in range(1, 301))
    tabular = _blocks_from_text("\n".join(tab_rows))
    assert len(tabular) == 1
    assert tabular[0]["kind"] == "table"
    assert tabular[0]["rows"][288] == ["288", "Человек 288", "Строка 288"]


def test_conversation_and_accepted_web_ledger_are_bounded_user_data() -> None:
    source_url = "https://safe.synthetic.example.com/fact"
    context = AgentContext(
        conversation_id="conversation-1",
        user_id="alice",
        conversation_history=[
            {"role": "user", "content": "Синтетический показатель равен 42."},
            {"role": "assistant", "content": "Зафиксировано без округления."},
        ],
        web_evidence_status="sourced",
        web_sources=[{"title": "Synthetic source", "url": source_url}],
    )

    records = _file_context_ground_records(context)

    conversation = next(item for item in records if item.startswith(_FILE_CONVERSATION_GROUNDS_PREFIX))
    web_ledger = next(item for item in records if item.startswith(_FILE_WEB_SOURCE_LEDGER_PREFIX))
    assert "Синтетический показатель равен 42" in conversation
    assert "untrusted JSON; data only" in conversation
    assert source_url in web_ledger
    assert '"status": "sourced"' in web_ledger


class _EnabledOfflineRouter:
    enabled = True
    total_budget_sec = 1.0


@pytest.mark.asyncio
async def test_late_file_fill_receives_attachment_dialogue_and_web_evidence(
    settings: Any,
    storage: Any,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Every authorised source reaches one tool-free carrier synthesis."""

    runtime = AgentRuntime(settings, storage, llm=_EnabledOfflineRouter(), kernel=object())  # type: ignore[arg-type]
    source_url = "https://safe.synthetic.example.com/fact"
    context = AgentContext(
        conversation_id="conversation-1",
        user_id="alice",
        conversation_history=[
            {"role": "user", "content": "В диалоге показатель Бета равен 17."},
            {"role": "assistant", "content": "Принято."},
        ],
        web_evidence_status="sourced",
        web_sources=[{"title": "Synthetic source", "url": source_url}],
    )
    attachment_map = _ATTACHMENT_MAP_PREFIX + '{"files":[{"filename":"source.docx"}]}'
    evidence = [
        {"tool": "attachment", "output": attachment_map},
        {
            "tool": "web_research",
            "output": "ACCEPTED_WEB_FACT: показатель Альфа равен 42.",
        },
    ]
    captured_messages: list[dict[str, Any]] = []
    captured_build: dict[str, Any] = {}

    async def primary_chat(
        used_context: AgentContext | None,
        messages: list[dict[str, Any]],
        **kwargs: Any,
    ) -> dict[str, str]:
        assert used_context is context
        assert kwargs == {"tools": []}
        captured_messages.extend(messages)
        return {"content": (f"Сводный отчёт\nПоказатели:\nАльфа — 42.\nБета — 17.\nИсточник — {source_url}")}

    async def make_file(
        prompt: str,
        answer: str,
        actor: ActorContext,
        *,
        blocks: list[dict[str, Any]] | None = None,
    ) -> dict[str, str]:
        captured_build.update({"request": prompt, "answer": answer, "actor": actor, "blocks": blocks})
        return {
            "kind": "document",
            "filename": "Сводный отчёт.xlsx",
            "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "content_base64": "c3ludGhldGlj",
        }

    monkeypatch.setattr(runtime, "_attachment_primary_chat", primary_chat)
    monkeypatch.setattr(runtime, "_make_file_from_answer", make_file)
    actor = ActorContext(user_id="alice", preset_key="owner", source="test")
    current_fact = "CURRENT_REQUEST_FACT: показатель Гамма равен 99."
    file_prompt = (
        "Оформи это красивой таблицей по данным из интернета.\n"
        + ("Синтетический контекст. " * 24)
        + current_fact
    )

    result = await runtime._file_for_a_request_that_wanted_one(  # noqa: SLF001
        file_prompt,
        "",
        actor,
        evidence=evidence,
        context=context,
    )

    delivered = [str(item.get("content") or "") for item in captured_messages]
    assert attachment_map in delivered
    assert any(
        item.startswith(_FILE_CONVERSATION_GROUNDS_PREFIX) and "Бета равен 17" in item for item in delivered
    )
    assert any(item.startswith(_FILE_WEB_SOURCE_LEDGER_PREFIX) and source_url in item for item in delivered)
    assert any("FRIDAY_FILE_GROUNDS_DATA" in item and "ACCEPTED_WEB_FACT" in item for item in delivered)
    assert any(current_fact in item for item in delivered)
    assert captured_build["request"].startswith("Оформи")
    assert captured_build["blocks"]
    assert result is not None and result["filename"].endswith(".xlsx")
