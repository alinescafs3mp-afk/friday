"""Strict CS1-gated bare-upload quicklook truth table.

Authority is stamped FileEvidenceView only. Public Mapping flags cannot mint
literals. Budgets and multi-file isolation are closed here.
"""

from __future__ import annotations

import asyncio
import io
import json
import zipfile
from dataclasses import replace
from typing import Any

import pytest
from openpyxl import Workbook

from friday.agent_runtime import (
    _BARE_UPLOAD_REVIEW_TASK,
    _FOCUSED_ATTACHMENT_CONTEXT_CHARS,
    _QUICKLOOK_MULTI_MAX_CHARS,
    _QUICKLOOK_MULTI_MAX_SNIPPETS,
    _QUICKLOOK_SINGLE_MAX_CHARS,
    _QUICKLOOK_SINGLE_MAX_SNIPPETS,
    _QUICKLOOK_TOTAL_MAX_CHARS,
    _QUICKLOOK_TRUNCATION_MARK,
    _UPLOAD_OVERVIEW_HEADING,
    AgentRuntime,
    FileBodyKind,
    FileRegistrationKind,
    _bounded_attachment_projection,
    _build_file_evidence_view,
    _file_evidence_set_from_attachments,
    _maybe_bounded_file_overview,
    _OwnedAttachment,
    _registered_upload_receipt_answer,
    _stamp_file_evidence,
    _upload_overview_set_admitted,
    _upload_overview_source_slices,
)
from friday.agent_runtime._office_attachments import (
    OFFICE_STRUCTURE_KEY,
    trusted_office_attachment,
)
from friday.ingestion import IngestionPipeline
from friday.knowledge_graph import KnowledgeGraph
from friday.permissions import ActorContext


class _NoQuicklookLLM:
    enabled = True
    model = "result20-quicklook"

    async def chat(self, messages, **kwargs):  # pragma: no cover - terminal route owns the turn
        del messages, kwargs
        raise AssertionError("bare upload quicklook called the model")


def _actor() -> ActorContext:
    return ActorContext(user_id="alice", preset_key="owner", source="test")


def _stamped_owned(
    *,
    raw_id: str,
    filename: str,
    text: str,
    record: str = "valid",
    disk: bool = True,
    advisory: bool = False,
    empty: bool = False,
    truncated: bool = False,
    verification_eligible: bool = True,
    stamp: bool = True,
) -> _OwnedAttachment:
    item = _OwnedAttachment(
        {
            "raw_object_id": raw_id,
            "filename": filename,
            "transient_text": text,
            "extraction_success": True,
            "verification_eligible": verification_eligible,
            "_registered_file_record": record,
            "_registered_file_bytes_verified": disk,
            "advisory_only": advisory,
            "empty_text": empty,
            "text_truncated": truncated,
            "extraction_truncated": truncated,
        }
    )
    if stamp:
        view = _build_file_evidence_view(item)
        assert view is not None
        _stamp_file_evidence(item, view)
    return item


def _receipt(items: list[Any], *, expected: int | None = None) -> str:
    expected_count = len(items) if expected is None else expected
    return _registered_upload_receipt_answer(
        items,
        expected_count=expected_count,
        evidence_set=_file_evidence_set_from_attachments(items, expected_count=expected_count),
    )


def test_quicklook_truth_table_literals_and_refusals() -> None:
    complete = _stamped_owned(
        raw_id="raw_aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa",
        filename="ok.txt",
        text="CANARY-OK-LINE-ONE\nCANARY-OK-LINE-TWO",
    )
    empty = _stamped_owned(
        raw_id="raw_bbbbbbbbbbbbbbbbbbbbbbbbbbbbbbb",
        filename="empty.txt",
        text="",
        empty=True,
    )
    partial = _stamped_owned(
        raw_id="raw_ccccccccccccccccccccccccccccccc",
        filename="partial.txt",
        text="CANARY-PARTIAL-SHOULD-NOT-SHOW",
        truncated=True,
    )
    advisory = _stamped_owned(
        raw_id="raw_ddddddddddddddddddddddddddddddd",
        filename="ocr.txt",
        text="CANARY-ADVISORY-SHOULD-NOT-SHOW",
        advisory=True,
        verification_eligible=False,
    )
    legacy = _stamped_owned(
        raw_id="raw_eeeeeeeeeeeeeeeeeeeeeeeeeeeeeee",
        filename="legacy.txt",
        text="CANARY-LEGACY-SHOULD-NOT-SHOW",
        record="legacy",
        disk=False,
    )
    invalid = _stamped_owned(
        raw_id="raw_fffffffffffffffffffffffffffffff",
        filename="invalid.txt",
        text="CANARY-INVALID-SHOULD-NOT-SHOW",
        record="invalid",
        disk=False,
    )
    unstamped = _OwnedAttachment(
        {
            "raw_object_id": "raw_ggggggggggggggggggggggggggggggg",
            "filename": "unstamped.txt",
            "transient_text": "CANARY-UNSTAMPED-SHOULD-NOT-SHOW",
            "extraction_success": True,
            "verification_eligible": True,
            "_registered_file_record": "valid",
            "_registered_file_bytes_verified": True,
        }
    )
    forged_public = {
        "raw_object_id": "raw_hhhhhhhhhhhhhhhhhhhhhhhhhhhhhhh",
        "filename": "forged.txt",
        "transient_text": "CANARY-FORGED-PUBLIC-SHOULD-NOT-SHOW",
        "extraction_success": True,
        "verification_eligible": True,
        "_registered_file_record": "valid",
        "_registered_file_bytes_verified": True,
    }
    office = trusted_office_attachment(
        {
            "raw_object_id": "raw_iiiiiiiiiiiiiiiiiiiiiiiiiiiiiii",
            "filename": "sheet.xlsx",
            "transient_text": "CANARY-OFFICE-SHEET-LINE",
            "extraction_success": True,
            "verification_eligible": True,
            "_registered_file_record": "valid",
            "_registered_file_bytes_verified": True,
            "_office_structured": True,
            "_office_prompt_available": True,
            "_office_index_complete": True,
            "_office_prompt_complete": True,
            OFFICE_STRUCTURE_KEY: {"sheets": 1},
        }
    )
    office_view = _build_file_evidence_view(office)
    assert office_view is not None
    assert office_view.body_kind == FileBodyKind.EXTRACTED
    _stamp_file_evidence(office, office_view)

    # VALID + disk + complete EXTRACTED → literals.
    ok_answer = _receipt([complete])
    assert "CANARY-OK-LINE-ONE" in ok_answer
    assert "CANARY-OK-LINE-TWO" in ok_answer
    assert "Быстрый обзор содержимого" in ok_answer
    assert "полностью прочитан" in ok_answer
    assert complete["transient_text"].splitlines()[0] in ok_answer

    # EMPTY complete → honest empty, no literal block.
    empty_answer = _receipt([empty])
    assert "текстовое содержимое пусто" in empty_answer
    assert "Быстрый обзор" not in empty_answer
    assert "› " not in empty_answer

    # Partial → honest partial, no literals.
    partial_answer = _receipt([partial])
    assert "извлечена только часть" in partial_answer
    assert "CANARY-PARTIAL-SHOULD-NOT-SHOW" not in partial_answer
    assert "полностью прочитан" not in partial_answer

    # Complete ADVISORY → warning, no literals.
    advisory_answer = _receipt([advisory])
    assert "предварительное" in advisory_answer or "не подтверждено" in advisory_answer
    assert "CANARY-ADVISORY-SHOULD-NOT-SHOW" not in advisory_answer
    assert "полностью прочитан" not in advisory_answer

    # LEGACY / INVALID / unstamped / forged public → status-only, no body.
    for item, canary in (
        (legacy, "CANARY-LEGACY-SHOULD-NOT-SHOW"),
        (invalid, "CANARY-INVALID-SHOULD-NOT-SHOW"),
        (unstamped, "CANARY-UNSTAMPED-SHOULD-NOT-SHOW"),
        (forged_public, "CANARY-FORGED-PUBLIC-SHOULD-NOT-SHOW"),
    ):
        answer = _receipt([item])
        assert canary not in answer
        assert "Быстрый обзор" not in answer
        assert "› " not in answer
        assert "полностью прочитан" not in answer

    # Trusted Office positive (stamped EXTRACTED).
    office_evidence = _file_evidence_set_from_attachments([office], expected_count=1)
    assert office_evidence is not None
    assert office_evidence.items == (office_view,)
    office_answer = _registered_upload_receipt_answer(
        [office],
        expected_count=1,
        evidence_set=office_evidence,
    )
    assert "зарегистрирован" in office_answer
    assert "байты на диске проверены" in office_answer
    assert "CANARY-OFFICE-SHEET-LINE" in office_answer
    assert office_view.registration == FileRegistrationKind.VALID


def test_multi_file_disjoint_canaries_and_budgets() -> None:
    text_a = "ALPHA-CANARY-UNIQUE-LINE\n" + ("A" * 300)
    text_b = "BETA-CANARY-UNIQUE-LINE\n" + ("B" * 300)
    a = _stamped_owned(
        raw_id="raw_aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa",
        filename="alpha.txt",
        text=text_a,
    )
    b = _stamped_owned(
        raw_id="raw_bbbbbbbbbbbbbbbbbbbbbbbbbbbbbbb",
        filename="beta.txt",
        text=text_b,
    )
    answer = _receipt([a, b])
    assert len(answer) <= _QUICKLOOK_TOTAL_MAX_CHARS
    assert "ALPHA-CANARY-UNIQUE-LINE" in answer
    assert "BETA-CANARY-UNIQUE-LINE" in answer
    # One block per file; no cross-label of body under wrong name.
    alpha_idx = answer.index("«alpha.txt»")
    beta_idx = answer.index("«beta.txt»")
    assert alpha_idx < beta_idx
    # Multi budget: at most one snippet of 160 chars per file.
    assert answer.count("› ") <= 2
    for line in answer.splitlines():
        if line.startswith("› "):
            payload = line[2:]
            assert len(payload) <= _QUICKLOOK_MULTI_MAX_CHARS
            assert payload in text_a or payload in text_b
    # Truncation marker is a separate service line, never ellipsis inside quote.
    if _QUICKLOOK_TRUNCATION_MARK in answer:
        assert "…" not in answer
    # Lead never claims full read for multi with any partial — both complete here.
    assert "полностью прочитан" not in answer or answer.startswith("Файл сохранён")
    # Multi lead is the multi form (not single-file full-read claim for the set).
    assert "Файлы зарегистрированы" in answer or "состояние чтения" in answer

    # Single-file long line: exact prefix + separate truncation mark.
    long_line = "Z" * 400
    single = _stamped_owned(
        raw_id="raw_ccccccccccccccccccccccccccccccc",
        filename="long.txt",
        text=long_line,
    )
    single_answer = _receipt([single])
    assert f"› {long_line[:_QUICKLOOK_SINGLE_MAX_CHARS]}" in single_answer
    assert _QUICKLOOK_TRUNCATION_MARK in single_answer
    assert "…" not in single_answer
    assert long_line[:_QUICKLOOK_SINGLE_MAX_CHARS] in long_line
    # Single allows up to 3 snippets.
    multi_line = "\n".join(f"LINE-{i}-MARKER" for i in range(6))
    many = _stamped_owned(
        raw_id="raw_ddddddddddddddddddddddddddddddd",
        filename="many.txt",
        text=multi_line,
    )
    many_answer = _receipt([many])
    assert many_answer.count("› ") == _QUICKLOOK_SINGLE_MAX_SNIPPETS
    assert "LINE-0-MARKER" in many_answer
    assert "LINE-3-MARKER" not in many_answer

    # Unsafe control characters are rejected (no normalized display).
    unsafe = _stamped_owned(
        raw_id="raw_eeeeeeeeeeeeeeeeeeeeeeeeeeeeeee",
        filename="unsafe.txt",
        text="SAFE-LINE\nBAD\x00CONTROL-LINE\nOTHER-SAFE",
    )
    unsafe_answer = _receipt([unsafe])
    assert "SAFE-LINE" in unsafe_answer
    assert "BAD" not in unsafe_answer or "\x00" not in unsafe_answer
    assert "CONTROL-LINE" not in unsafe_answer
    assert "OTHER-SAFE" in unsafe_answer

    # Budget constants stay closed for multi.
    assert _QUICKLOOK_MULTI_MAX_SNIPPETS == 1
    assert _QUICKLOOK_MULTI_MAX_CHARS == 160


def test_status_only_twelve_long_names_obey_absolute_total_budget() -> None:
    items = [
        _OwnedAttachment(
            {
                "raw_object_id": f"raw_result20_status_{index:02d}",
                "filename": f"RESULT20-STATUS-{index:02d}-" + ("N" * 245),
                "transient_text": f"RESULT20-STATUS-BODY-MUST-NOT-LEAK-{index:02d}",
                "extraction_success": True,
                "verification_eligible": True,
                "_registered_file_record": "valid",
                "_registered_file_bytes_verified": True,
            }
        )
        for index in range(12)
    ]

    answer = _receipt(items, expected=12)

    assert len(answer) <= _QUICKLOOK_TOTAL_MAX_CHARS
    assert answer.count("\n• ") == 12
    positions = [answer.index(f"RESULT20-STATUS-{index:02d}-") for index in range(12)]
    assert positions == sorted(positions)
    assert answer.count("[имя сокращено]") == 12
    assert "RESULT20-STATUS-BODY-MUST-NOT-LEAK" not in answer
    assert "Быстрый обзор" not in answer


@pytest.mark.asyncio
async def test_bare_upload_uses_normal_hierarchy_when_prompt_projection_is_truncated(
    settings,
    storage,
    monkeypatch,
) -> None:
    configured = replace(settings, verify_answers=False)
    storage.ensure_user("alice", preset_key="owner")
    source = "RESULT20-LONG-ACTIVE-EXACT-LITERAL\n" + (
        "registered active source remains complete\n" * ((_FOCUSED_ATTACHMENT_CONTEXT_CHARS // 42) + 200)
    )
    assert len(source) > _FOCUSED_ATTACHMENT_CONTEXT_CHARS
    pipeline = IngestionPipeline(configured, storage, KnowledgeGraph(storage))
    ingested = await pipeline.ingest_file(
        "alice",
        None,
        source.encode(),
        filename="result20-long-complete.txt",
        mime_type="text/plain",
        metadata={"uploaded_by": "alice"},
        source_ref="telegram-file:RESULT20-LONG-ACTIVE",
    )
    runtime = AgentRuntime(configured, storage, llm=_NoQuicklookLLM())  # type: ignore[arg-type]
    calls = {"hierarchy": 0, "answer": 0}

    async def hierarchy(context, message, attachments, *, task_kind):  # noqa: ANN001
        del context
        calls["hierarchy"] += 1
        assert task_kind == "summary"
        assert message == _BARE_UPLOAD_REVIEW_TASK
        assert len(attachments) == 1
        assert str(attachments[0].get("transient_text") or "") == source
        return None, False

    async def answer(context, message, attachments, **kwargs):  # noqa: ANN001
        del context, kwargs
        calls["answer"] += 1
        assert message == _BARE_UPLOAD_REVIEW_TASK
        assert len(attachments) == 1
        assert str(attachments[0].get("transient_text") or "") == source
        return {
            "content": "Подробное ревью большого файла: контрольный литерал подтверждён.",
            "tools_used": [],
            "_model_generated": True,
        }

    async def forbidden_quicklook(*args, **kwargs):  # noqa: ANN002, ANN003
        del args, kwargs
        raise AssertionError("bare upload entered the removed 20-second overview path")

    monkeypatch.setattr(runtime, "_build_attachment_hierarchy_bundle", hierarchy)
    monkeypatch.setattr(runtime, "_hierarchical_attachment_response", answer)
    monkeypatch.setattr("friday.agent_runtime._maybe_bounded_file_overview", forbidden_quicklook)
    receipt = await runtime.chat(
        "alice",
        "Загружен документ: result20-long-complete.txt",
        actor=_actor(),
        attachments=[{"raw_object_id": str(ingested["raw_object_id"])}],
        synthetic_document_notice=True,
    )

    assert receipt["message"].endswith("Подробное ревью большого файла: контрольный литерал подтверждён.")
    assert "Не весь исходный материал" in receipt["message"]
    assert "Быстрый обзор" not in receipt["message"]
    assert receipt["message_format"] == "markdown"
    assert receipt["tools_used"] == []
    assert calls == {"hierarchy": 1, "answer": 1}


@pytest.mark.asyncio
async def test_nine_complete_office_uploads_use_hierarchy_when_prompt_projection_reads_only_two(
    settings,
    storage,
    monkeypatch,
) -> None:
    """Prompt capacity is not source readability for a complete nine-file set."""

    configured = replace(settings, verify_answers=False)
    storage.ensure_user("alice", preset_key="owner")
    pipeline = IngestionPipeline(configured, storage, KnowledgeGraph(storage))
    raw_ids: list[str] = []
    source_texts: list[str] = []
    # File one fits, file two partially fills the 24k Office envelope, and the
    # remaining seven are content-free in that *prompt projection*.  All nine
    # registered sources themselves remain fully parsed and readable.
    row_counts = [2, 80, 100, 100, 100, 100, 100, 100, 100]
    for file_index, row_count in enumerate(row_counts, start=1):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = f"SOURCE-{file_index}"
        sheet.append(["№", "Значение", "Описание"])
        for row_index in range(1, row_count + 1):
            sheet.append(
                [
                    row_index,
                    f"FILE-{file_index}-VALUE-{row_index}",
                    f"FILE-{file_index}-ROW-{row_index}-" + ("данные " * 12),
                ]
            )
        stream = io.BytesIO()
        workbook.save(stream)
        workbook.close()
        outcome = await pipeline.ingest_file(
            "alice",
            None,
            stream.getvalue(),
            filename=f"projection-heavy-{file_index}.xlsx",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            metadata={"uploaded_by": "alice"},
            source_ref=f"telegram-file:RESULT20-NINE-{file_index}",
        )
        raw_id = str(outcome["raw_object_id"])
        raw_ids.append(raw_id)
        raw = storage.get_raw_object(raw_id, "alice")
        assert raw is not None
        source_texts.append(str(raw["raw_content"]))

    assert sum(map(len, source_texts)) > _FOCUSED_ATTACHMENT_CONTEXT_CHARS
    runtime = AgentRuntime(configured, storage, llm=_NoQuicklookLLM())  # type: ignore[arg-type]
    owned = [
        runtime._owned_file_attachment(  # noqa: SLF001
            raw_id,
            tenant_id="alice",
            person_id="alice",
        )
        for raw_id in raw_ids
    ]
    assert all(item is not None for item in owned)
    verified = await runtime._verify_registered_file_attachments(  # noqa: SLF001
        [item for item in owned if item is not None],
        tenant_id="alice",
        person_id="alice",
    )
    projection = _bounded_attachment_projection(verified)
    projected_set = _file_evidence_set_from_attachments(projection, expected_count=9)
    source_set = _file_evidence_set_from_attachments(verified, expected_count=9)
    assert source_set is not None
    assert source_set.source_readable_count == 9
    assert source_set.coverage_complete is True
    assert source_set.verification_complete is True
    assert projected_set is not None
    assert projected_set.source_readable_count == 2
    assert projected_set.coverage_complete is False

    calls = {"hierarchy": 0, "answer": 0}

    async def hierarchy(context, message, attachments, *, task_kind):  # noqa: ANN001
        del context
        calls["hierarchy"] += 1
        assert message == _BARE_UPLOAD_REVIEW_TASK
        assert task_kind == "summary"
        assert len(attachments) == 9
        assert [str(item.get("transient_text") or "") for item in attachments] == source_texts
        return None, False

    async def answer(context, message, attachments, **kwargs):  # noqa: ANN001
        del context, kwargs
        calls["answer"] += 1
        assert message == _BARE_UPLOAD_REVIEW_TASK
        assert len(attachments) == 9
        return {
            "content": "Содержательное ревью полного набора из девяти источников.",
            "tools_used": [],
            "_model_generated": True,
        }

    monkeypatch.setattr(runtime, "_build_attachment_hierarchy_bundle", hierarchy)
    monkeypatch.setattr(runtime, "_hierarchical_attachment_response", answer)
    receipt = await runtime.chat(
        "alice",
        "Загружено документов: 9",
        actor=_actor(),
        attachments=[{"raw_object_id": raw_id} for raw_id in raw_ids],
        synthetic_document_notice=True,
    )

    assert calls == {"hierarchy": 1, "answer": 1}
    assert "Содержательное ревью полного набора" in receipt["message"]
    assert "доступно для 2 из 9" not in receipt["message"]

    corrupt = await pipeline.ingest_file(
        "alice",
        None,
        b"not-an-xlsx-container",
        filename="unreadable-9.xlsx",
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        metadata={"uploaded_by": "alice"},
        source_ref="telegram-file:RESULT20-NINE-UNREADABLE",
    )
    corrupt_raw_id = str(corrupt["raw_object_id"])
    corrupt_owned = runtime._owned_file_attachment(  # noqa: SLF001
        corrupt_raw_id,
        tenant_id="alice",
        person_id="alice",
    )
    assert corrupt_owned is not None
    corrupt_verified = await runtime._verify_registered_file_attachments(  # noqa: SLF001
        [corrupt_owned],
        tenant_id="alice",
        person_id="alice",
    )
    assert len(corrupt_verified) == 1
    mixed_sources = [*verified[:-1], corrupt_verified[0]]
    mixed_projection = _bounded_attachment_projection(mixed_sources)
    mixed_projected_set = _file_evidence_set_from_attachments(mixed_projection, expected_count=9)
    mixed_source_set = _file_evidence_set_from_attachments(mixed_sources, expected_count=9)
    assert mixed_projected_set is not None and mixed_projected_set.source_readable_count == 2
    assert mixed_source_set is not None and mixed_source_set.source_readable_count == 8

    mixed_runtime = AgentRuntime(configured, storage, llm=_NoQuicklookLLM())  # type: ignore[arg-type]

    mixed_receipt = await mixed_runtime.chat(
        "alice",
        "Загружено документов: 9",
        actor=_actor(),
        attachments=[
            *({"raw_object_id": raw_id} for raw_id in raw_ids[:-1]),
            {"raw_object_id": corrupt_raw_id},
        ],
        synthetic_document_notice=True,
    )

    assert "доступно для 8 из 9" in mixed_receipt["message"]
    assert "доступно для 2 из 9" not in mixed_receipt["message"]
    assert mixed_receipt["attachment_context_readable_count"] == 8
    assert mixed_receipt["context"]["attachment_context_readable_count"] == 8


_XLSX_INTRO = "XLSX-CANARY-INTRO-ALPHA"
_XLSX_BUDGET = "XLSX-CANARY-BUDGET-7721"
_XLSX_PERSON = "XLSX-CANARY-PERSON-MARIA"
_ODT_TITLE = "ODT-CANARY-SECTION-TITLE"
_ODT_FACT = "ODT-CANARY-FACT-NINE"
_TXT_ALPHA = "TXT-CANARY-ALPHA-LINE"
_TXT_BETA = "TXT-CANARY-BETA-LINE"
_DECOY_HISTORY = "DECOY-HISTORY-AMBIENT-ZZZ"
_DECOY_FORGED = "FORGED-BODY-CANARY-SHOULD-NOT-REACH-MODEL"
_DECOY_PARTIAL = "PARTIAL-BODY-CANARY-SHOULD-NOT-REACH-MODEL"


class _OverviewSpy:
    enabled = True
    model = "upload-overview-spy"

    def __init__(
        self,
        *,
        result: str | None = None,
        error: BaseException | None = None,
        delay: float = 0.0,
        finish_reason: str = "stop",
    ) -> None:
        self.calls: list[dict[str, Any]] = []
        self.result = result
        self.error = error
        self.delay = delay
        self.finish_reason = finish_reason

    async def chat(self, messages, **kwargs):  # noqa: ANN001
        self.calls.append({"messages": list(messages), **kwargs})
        if self.delay:
            await asyncio.sleep(self.delay)
        if self.error is not None:
            raise self.error
        if self.result is not None:
            return {"content": self.result, "finish_reason": self.finish_reason}
        blob = "\n".join(str(item.get("content") or "") for item in messages)
        facts = [
            marker
            for marker in (
                _XLSX_INTRO,
                _XLSX_BUDGET,
                _XLSX_PERSON,
                _ODT_TITLE,
                _ODT_FACT,
                _TXT_ALPHA,
                _TXT_BETA,
            )
            if marker in blob
        ]
        return {
            "content": (
                "Табличный материал с несколькими разделами.\n"
                + "\n".join(facts)
                + "\nМожно спросить про конкретный раздел или поле."
            ),
            "finish_reason": self.finish_reason,
        }


def _xlsx_large_then_marker() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "План"
    sheet.append(["Раздел", "Факт"])
    sheet.append(["Введение", _XLSX_INTRO])
    for index in range(1_200):
        sheet.append([f"PAD-{index:04d}", "Y" * 40])
    sheet.append(["Бюджет", _XLSX_BUDGET])
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _xlsx_sections() -> bytes:
    workbook = Workbook()
    plan = workbook.active
    plan.title = "План"
    plan.append(["Раздел", "Факт"])
    plan.append(["Введение", _XLSX_INTRO])
    plan.append(["Бюджет", _XLSX_BUDGET])
    people = workbook.create_sheet("Люди")
    people.append(["ФИО", "Роль"])
    people.append([_XLSX_PERSON, "аналитик"])
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _odt_sections() -> bytes:
    payload = io.BytesIO()
    with zipfile.ZipFile(payload, "w") as archive:
        archive.writestr("mimetype", "application/vnd.oasis.opendocument.text")
        archive.writestr(
            "content.xml",
            """<?xml version="1.0" encoding="UTF-8"?>
<office:document-content
 xmlns:office="urn:oasis:names:tc:opendocument:xmlns:office:1.0"
 xmlns:text="urn:oasis:names:tc:opendocument:xmlns:text:1.0">
 <office:body><office:text>"""
            + f"<text:p>{_ODT_TITLE}</text:p><text:p>{_ODT_FACT}</text:p>"
            + """</office:text></office:body>
</office:document-content>""",
        )
        archive.writestr(
            "meta.xml",
            """<?xml version="1.0" encoding="UTF-8"?>
<office:document-meta
 xmlns:office="urn:oasis:names:tc:opendocument:xmlns:office:1.0"
 xmlns:dc="http://purl.org/dc/elements/1.1/">
 <office:meta><dc:title>ODT overview</dc:title></office:meta>
</office:document-meta>""",
        )
    return payload.getvalue()


def _prompt_blob(spy: _OverviewSpy) -> str:
    return "\n".join(
        str(item.get("content") or "")
        for call in spy.calls
        for item in call["messages"]
        if isinstance(item, dict)
    )


def _assert_public_file_metrics(receipt: dict[str, Any], *, expected: int) -> None:
    assert receipt["attachment_context_expected_count"] == expected
    assert receipt["attachment_context_readable_count"] == expected
    assert receipt["attachment_coverage_complete"] is True
    assert receipt["attachment_verification_complete"] is True
    assert receipt["tools_used"] == []
    assert receipt["message_format"] == "markdown"


async def _ingest(settings, storage, payload: bytes, filename: str, mime_type: str) -> str:
    pipeline = IngestionPipeline(settings, storage, KnowledgeGraph(storage))
    ingested = await pipeline.ingest_file(
        "alice",
        None,
        payload,
        filename=filename,
        mime_type=mime_type,
        metadata={"uploaded_by": "alice"},
        source_ref=f"telegram-file:{filename}",
    )
    return str(ingested["raw_object_id"])


async def _upload_turn(
    runtime: AgentRuntime,
    attachments: list[dict[str, Any]],
    *,
    filename: str,
    conversation_id: str | None = None,
) -> dict[str, Any]:
    return await runtime.chat(
        "alice",
        f"Загружен документ: {filename}",
        actor=_actor(),
        conversation_id=conversation_id,
        attachments=attachments,
        synthetic_document_notice=True,
    )


def _forbid_generic(name: str):
    async def forbidden(*args, **kwargs):  # noqa: ANN002, ANN003
        del args, kwargs
        raise AssertionError(f"bounded overview entered {name}")

    return forbidden


@pytest.mark.asyncio
async def test_registered_xlsx_upload_and_adjacent_overview(settings, storage) -> None:
    configured = replace(settings, verify_answers=False)
    storage.ensure_user("alice", preset_key="owner")
    conversation = storage.create_conversation("alice", title="overview-decoy")
    storage.store_message(str(conversation["id"]), "alice", "user", _DECOY_HISTORY)
    storage.store_message(str(conversation["id"]), "alice", "assistant", "DECOY-KO-AMBIENT-QQQ")
    xlsx_id = await _ingest(
        configured,
        storage,
        _xlsx_sections(),
        "plan-people.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    other_id = await _ingest(
        configured,
        storage,
        f"{_TXT_ALPHA}\nreply target\n".encode(),
        "reply-a.txt",
        "text/plain",
    )
    spy = _OverviewSpy()
    runtime = AgentRuntime(configured, storage, llm=spy)  # type: ignore[arg-type]
    receipt = await _upload_turn(
        runtime,
        [{"raw_object_id": xlsx_id}],
        filename="plan-people.xlsx",
        conversation_id=str(conversation["id"]),
    )

    assert "Файл сохранён" not in receipt["message"]
    assert _UPLOAD_OVERVIEW_HEADING not in receipt["message"]
    assert _XLSX_INTRO in receipt["message"]
    assert _XLSX_BUDGET in receipt["message"]
    assert _XLSX_PERSON in receipt["message"]
    assert _DECOY_HISTORY not in receipt["message"]
    _assert_public_file_metrics(receipt, expected=1)
    assert len(spy.calls) == 1
    assert spy.calls[0].get("tools") == []
    prompt = _prompt_blob(spy)
    assert "содержательное подробное ревью" in prompt
    assert _XLSX_INTRO in prompt
    assert _XLSX_PERSON in prompt
    assert _DECOY_HISTORY not in prompt
    assert "DECOY-KO-AMBIENT-QQQ" not in prompt
    meta = json.loads(str(storage.get_message(str(receipt["message_id"]), "alice")["metadata_json"] or "{}"))
    assert meta.get("overview_model_used") is not True
    assert meta.get("structural", {}).get("model_spoke") is True

    adjacent = await runtime.chat(
        "alice",
        "дай обзор файла",
        actor=_actor(),
        conversation_id=str(conversation["id"]),
        attachments=[],
    )
    assert _UPLOAD_OVERVIEW_HEADING not in adjacent["message"]
    assert _XLSX_INTRO in adjacent["message"]
    _assert_public_file_metrics(adjacent, expected=1)
    assert len(spy.calls) == 2
    assert xlsx_id in _prompt_blob(spy) or _XLSX_INTRO in _prompt_blob(spy)

    reply = await runtime.chat(
        "alice",
        "Кратко по файлу",
        actor=_actor(),
        conversation_id=str(conversation["id"]),
        attachments=[{"raw_object_id": other_id}],
        quoted_attachment_reference=True,
    )
    assert _TXT_ALPHA in reply["message"]
    last_prompt = "\n".join(str(item.get("content") or "") for item in spy.calls[-1]["messages"])
    assert _TXT_ALPHA in last_prompt
    _assert_public_file_metrics(reply, expected=1)

    voice_spy = _OverviewSpy()
    voice_runtime = AgentRuntime(configured, storage, llm=voice_spy)  # type: ignore[arg-type]
    hostile_filename_receipt = await _upload_turn(
        voice_runtime,
        [{"raw_object_id": xlsx_id}],
        filename="озвучь.txt",
    )
    assert _UPLOAD_OVERVIEW_HEADING not in hostile_filename_receipt["message"]
    assert hostile_filename_receipt["voice"] is None
    assert hostile_filename_receipt["tools_used"] == []
    assert len(voice_spy.calls) == 1

    # A backend-authored filename is inert data.  In particular, words inside
    # ``ВОТ ЭТОТ ФАЙЛ.docx`` cannot fabricate a current+prior comparison and
    # inflate one uploaded Raw into an expected set of two.
    count_spy = _OverviewSpy()
    count_runtime = AgentRuntime(configured, storage, llm=count_spy)  # type: ignore[arg-type]
    inert_name_receipt = await _upload_turn(
        count_runtime,
        [{"raw_object_id": xlsx_id}],
        filename="ВОТ ЭТОТ ФАЙЛ.docx",
    )
    _assert_public_file_metrics(inert_name_receipt, expected=1)
    assert inert_name_receipt["restored_attachment_count"] == 0
    inert_rows = storage.get_conversation_messages(
        str(inert_name_receipt["conversation_id"]),
        user_id="alice",
    )
    inert_user = next(item for item in inert_rows if item.get("role") == "user")
    inert_metadata = json.loads(str(inert_user.get("metadata_json") or "{}"))
    assert inert_metadata["attachment_count"] == 1

    empty_conv = storage.create_conversation("alice", title="overview-empty")
    empty_spy = _OverviewSpy()
    empty_runtime = AgentRuntime(configured, storage, llm=empty_spy)  # type: ignore[arg-type]
    closed = await empty_runtime.chat(
        "alice",
        "дай обзор файла",
        actor=_actor(),
        conversation_id=str(empty_conv["id"]),
        attachments=[],
    )
    assert closed["message"] == "Не могу сделать обзор: нет однозначно выбранного файла."
    assert closed["attachment_context_expected_count"] == 0
    assert closed["attachment_context_readable_count"] == 0
    assert closed["attachment_coverage_complete"] is False
    assert empty_spy.calls == []

    multi_conv = storage.create_conversation("alice", title="overview-multi")
    multi_runtime = AgentRuntime(configured, storage, llm=empty_spy)  # type: ignore[arg-type]
    await _upload_turn(
        multi_runtime,
        [{"raw_object_id": xlsx_id}, {"raw_object_id": other_id}],
        filename="plan-people.xlsx",
        conversation_id=str(multi_conv["id"]),
    )
    empty_spy.calls.clear()
    singular = await multi_runtime.chat(
        "alice",
        "дай обзор файла",
        actor=_actor(),
        conversation_id=str(multi_conv["id"]),
        attachments=[],
    )
    assert singular["message"] == "Не могу сделать обзор: нет однозначно выбранного файла."
    assert singular["attachment_context_expected_count"] == 0
    assert empty_spy.calls == []

    # The immediately preceding upload and the assistant's actually used
    # lineage must be the same ordered set. A user row that uploaded B while
    # explicitly selecting A cannot make a later shorthand silently switch to B.
    mismatch_conv = storage.create_conversation("alice", title="overview-lineage-mismatch")
    storage.store_message(
        str(mismatch_conv["id"]),
        "alice",
        "user",
        "Загружен B, выбран A",
        metadata={
            "conversation_uploaded_attachment_raw_ids": [other_id],
            "conversation_attachment_raw_ids": [xlsx_id],
        },
    )
    storage.store_message(
        str(mismatch_conv["id"]),
        "alice",
        "assistant",
        "Ответ по A",
        metadata={
            "attachment_context_used": True,
            "conversation_attachment_raw_ids": [xlsx_id],
        },
    )
    empty_spy.calls.clear()
    mismatch = await empty_runtime.chat(
        "alice",
        "дай обзор файла",
        actor=_actor(),
        conversation_id=str(mismatch_conv["id"]),
        attachments=[],
    )
    assert mismatch["message"] == "Не могу сделать обзор: нет однозначно выбранного файла."
    assert mismatch["attachment_context_expected_count"] == 0
    assert mismatch["attachment_context_readable_count"] == 0
    assert _XLSX_INTRO not in mismatch["message"]
    assert _TXT_ALPHA not in mismatch["message"]
    assert empty_spy.calls == []


@pytest.mark.asyncio
async def test_retired_bounded_overview_helper_still_fails_closed_on_partial_set() -> None:
    """Compatibility-only helper is not routed; if called directly it leaks no partial body."""

    complete = _stamped_owned(
        raw_id="raw_aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa",
        filename="alpha.txt",
        text=f"{_TXT_ALPHA}\nsection one body",
    )
    partial = _stamped_owned(
        raw_id="raw_bbbbbbbbbbbbbbbbbbbbbbbbbbbbbbb",
        filename="partial.txt",
        text=_DECOY_PARTIAL,
        truncated=True,
    )
    mixed_evidence = _file_evidence_set_from_attachments([complete, partial], expected_count=2)
    assert mixed_evidence is not None
    assert _upload_overview_set_admitted([complete, partial], evidence_set=mixed_evidence) is False
    assert _upload_overview_source_slices([complete, partial], evidence_set=mixed_evidence) == []
    partial_spy = _OverviewSpy()
    answer, used = await _maybe_bounded_file_overview(
        partial_spy,
        "Файлы зарегистрированы; состояние чтения указано ниже.",
        [complete, partial],
        evidence_set=mixed_evidence,
    )
    assert used is False
    assert answer == "Файлы зарегистрированы; состояние чтения указано ниже."
    assert partial_spy.calls == []
    assert _DECOY_PARTIAL not in answer
