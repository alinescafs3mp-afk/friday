"""K19 acceptance: every unambiguous ordinary table gets the exact path.

The parser may still fail closed on ambiguous layout.  What it may not do is
reserve authoritative row sets for tables whose header happens to name people.
All workbook/document values are generated from the frozen synthetic fixture.
"""

from __future__ import annotations

import copy
import io
import json
from pathlib import Path
from typing import Any

import pytest
from docx import Document
from openpyxl import Workbook

from friday.agent_runtime import _bounded_attachment_projection
from friday.agent_runtime._office_attachments import (
    OFFICE_STRUCTURE_KEY,
    code_owned_office_answer,
    trusted_office_attachment,
    validate_runtime_office_index,
)
from friday.documents import DocumentExtractor, validate_office_structure_index

FIXTURE_PATH = Path(__file__).parent / "fixtures" / "package_c_document_holdout.json"


def _fixture() -> dict[str, Any]:
    return json.loads(FIXTURE_PATH.read_text(encoding="utf-8"))


def _xlsx(rows: list[list[str]], *, merge: str = "", sibling_rows: list[list[str]] | None = None) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SYNTHETIC-MAIN"
    for row in rows:
        sheet.append(row)
    if merge:
        sheet.merge_cells(merge)
    if sibling_rows:
        sibling = workbook.create_sheet("SYNTHETIC-SIBLING")
        for row in sibling_rows:
            sibling.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _docx(rows: list[list[str]]) -> bytes:
    document = Document()
    width = max(len(row) for row in rows)
    table = document.add_table(rows=len(rows), cols=width)
    for row_number, values in enumerate(rows):
        for column_number, value in enumerate(values):
            table.cell(row_number, column_number).text = value
    buffer = io.BytesIO()
    document.save(buffer)
    return buffer.getvalue()


def _csv(rows: list[list[str]]) -> bytes:
    assert all(";" not in value and "\n" not in value for row in rows for value in row)
    return "\n".join(";".join(row) for row in rows).encode("utf-8")


def _extract(case: dict[str, Any]):  # noqa: ANN202 - DocumentResult is an internal carrier
    rows = [[str(value) for value in row] for row in case["rows"]]
    format_name = str(case["format"])
    if format_name == "xlsx":
        payload = _xlsx(
            rows,
            merge=str(case.get("merge") or ""),
            sibling_rows=case.get("sibling_rows"),
        )
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif format_name == "docx":
        payload = _docx(rows)
        mime = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    elif format_name == "csv":
        payload = _csv(rows)
        mime = "text/csv"
    else:  # pragma: no cover - fixture enum/shape test owns this branch
        raise AssertionError(format_name)
    result = DocumentExtractor(secret_values=()).extract(payload, str(case["filename"]), mime)
    assert result.success is True
    assert isinstance(result.office_structure_index, dict)
    assert (
        validate_office_structure_index(result.office_structure_index, result.text)
        == result.office_structure_index
    )
    return result


def _project(case: dict[str, Any]):  # noqa: ANN202 - private runtime carrier
    result = _extract(case)
    attachment = trusted_office_attachment(
        {
            "filename": case["filename"],
            "transient_text": result.text,
            "extraction_success": True,
            "verification_eligible": True,
            OFFICE_STRUCTURE_KEY: result.office_structure_index,
        }
    )
    return result, _bounded_attachment_projection([attachment])


def _record_literals(index: dict[str, Any], text: str, record_ids: list[str]) -> list[list[str]]:
    rows = {str(row["id"]): row for block in index["blocks"] for row in block.get("rows", [])}
    return [
        [text[cell["text_span"][0] : cell["text_span"][1]] for cell in rows[record_id]["cells"]]
        for record_id in record_ids
    ]


@pytest.mark.parametrize("case", _fixture()["k19_ordinary_tables"], ids=lambda case: case["id"])
def test_k19_unambiguous_ordinary_table_has_one_authoritative_exact_record_set(case: dict[str, Any]) -> None:
    result, projected = _project(case)
    index = result.office_structure_index or {}
    expected = case["expected"]

    assert index["complete"] is True
    assert len(index["record_sets"]) == 1, "ordinary table still has no authoritative rows"
    record_set = index["record_sets"][0]
    assert record_set["kind"] == "table_rows"
    assert record_set["authoritative"] is True
    assert record_set["person_column"] is None
    assert record_set["records_total"] == expected["records_total"]
    assert len(record_set["record_ids"]) == expected["records_total"]
    assert index["candidate_refs"] == []

    rows = [row for block in index["blocks"] for row in block.get("rows", [])]
    assert [row["role"] for row in rows if row["role"] != "empty"] == [
        "header",
        *(["record"] * expected["records_total"]),
    ]
    assert _record_literals(index, result.text, record_set["record_ids"]) == case["rows"][1:]

    answer = code_owned_office_answer(
        case["question"],
        projected,
        kind_override=expected["exact_kind"],
    )
    assert answer is not None
    assert answer["status"] == expected["status"] == "passed"
    assert answer["kind"] == expected["exact_kind"]
    assert str(expected["records_total"]) in answer["content"]
    if expected["exact_kind"] == "list_records":
        for value in (value for row in case["rows"][1:] for value in row if value):
            assert value in answer["content"]


def test_k19_the_same_ordinary_rows_have_the_same_semantics_in_csv_xlsx_and_docx() -> None:
    seed = copy.deepcopy(_fixture()["k19_ordinary_tables"][0])
    observed: list[tuple[list[str], str, int, None, list[list[str]]]] = []
    for format_name in ("csv", "xlsx", "docx"):
        case = {
            **seed,
            "format": format_name,
            "filename": f"synthetic-parity.{format_name}",
        }
        result = _extract(case)
        index = result.office_structure_index or {}
        assert len(index["record_sets"]) == 1
        record_set = index["record_sets"][0]
        observed.append(
            (
                [row["role"] for block in index["blocks"] for row in block.get("rows", [])],
                str(record_set["kind"]),
                int(record_set["records_total"]),
                record_set["person_column"],
                _record_literals(index, result.text, list(record_set["record_ids"])),
            )
        )

    assert observed[0] == observed[1] == observed[2]
    assert observed[0] == (
        ["header", "record", "record", "record"],
        "table_rows",
        3,
        None,
        seed["rows"][1:],
    )


def test_k19_people_table_keeps_person_semantics_and_ordinary_table_cannot_invent_people() -> None:
    fixture = _fixture()
    people_case = fixture["k19_controls"][0]
    people_result, people_projection = _project(people_case)
    people_set = (people_result.office_structure_index or {})["record_sets"][0]

    assert people_set["kind"] == "person_rows"
    assert people_set["person_column"] == 1
    people_answer = code_owned_office_answer(
        people_case["question"],
        people_projection,
        kind_override=people_case["expected"]["exact_kind"],
    )
    assert people_answer is not None and people_answer["status"] == "passed"
    assert people_answer["kind"] == "count_people"

    ordinary_case = fixture["k19_ordinary_tables"][0]
    _, ordinary_projection = _project(ordinary_case)
    refused = code_owned_office_answer(
        "Перечисли всех людей из этого файла.",
        ordinary_projection,
        kind_override="list_people",
    )
    assert refused is not None and refused["status"] == "unknown"
    assert "SYN-ROW-01" not in refused["content"]


@pytest.mark.parametrize("case", _fixture()["k19_controls"][1:], ids=lambda case: case["id"])
def test_k19_ambiguous_or_uncovered_ordinary_regions_fail_closed(case: dict[str, Any]) -> None:
    _, projected = _project(case)

    answer = code_owned_office_answer(
        case["question"],
        projected,
        kind_override=case["expected"]["exact_kind"],
    )
    assert answer is not None
    assert answer["status"] == "unknown"
    assert answer["kind"] == "unavailable"
    for value in (
        value
        for row in [*case["rows"], *case.get("sibling_rows", [])]
        for value in row
        if value.startswith("SYN-") or value.startswith("SYNTHETIC-")
    ):
        assert value not in answer["content"]


def test_k19_coordinated_omission_cannot_turn_n_minus_one_rows_into_a_valid_exact_set() -> None:
    case = _fixture()["k19_ordinary_tables"][0]
    result = _extract(case)
    index = copy.deepcopy(result.office_structure_index or {})
    assert len(index["record_sets"]) == 1
    record_set = index["record_sets"][0]
    assert record_set["kind"] == "table_rows"

    omitted_id = record_set["record_ids"].pop()
    record_set["records_total"] -= 1
    omitted_row = next(
        row for block in index["blocks"] for row in block.get("rows", []) if row["id"] == omitted_id
    )
    omitted_row["role"] = "unknown"

    assert validate_office_structure_index(index, result.text) is None
    assert validate_runtime_office_index(index, result.text) is None


def test_k19_an_ordinary_set_cannot_be_mutated_into_a_people_set_by_adding_a_column() -> None:
    case = _fixture()["k19_ordinary_tables"][1]
    result = _extract(case)
    index = copy.deepcopy(result.office_structure_index or {})
    assert len(index["record_sets"]) == 1
    index["record_sets"][0]["person_column"] = 1

    assert validate_office_structure_index(index, result.text) is None


@pytest.mark.parametrize(
    ("case_id", "rows"),
    [
        pytest.param(
            "annotated-repeated-header",
            [
                ["ID", "Статус"],
                ["SYN-ROW-01", "готово"],
                ["ID", "Статус (продолжение)"],
                ["SYN-ROW-02", "готово"],
            ],
            id="annotated-repeated-header",
        ),
        pytest.param(
            "unknown-note-footer",
            [
                ["ID", "Статус"],
                ["SYN-ROW-01", "готово"],
                ["Примечание", "SYNTHETIC-NOT-A-RECORD"],
            ],
            id="unknown-note-footer",
        ),
        pytest.param(
            "terminal-numeric-total",
            [
                ["ID", "Статус"],
                ["SYN-ROW-01", "готово"],
                ["SYN-ROW-02", "готово"],
                ["", "2"],
            ],
            id="terminal-numeric-total",
        ),
        pytest.param(
            "source-footer",
            [
                ["ID", "Статус"],
                ["SYN-ROW-01", "готово"],
                ["SYN-ROW-02", "готово"],
                ["Источник", "SYNTHETIC-REFERENCE-NOT-A-RECORD"],
            ],
            id="source-footer",
        ),
        pytest.param(
            "dated-annotation-footer",
            [
                ["ID", "Статус"],
                ["SYN-ROW-01", "готово"],
                ["SYN-ROW-02", "готово"],
                ["Data as of 2026-08-08", "SYNTHETIC-REFERENCE-NOT-A-RECORD"],
            ],
            id="dated-annotation-footer",
        ),
        pytest.param(
            "prepared-by-footer",
            [
                ["ID", "Статус"],
                ["SYN-ROW-01", "готово"],
                ["SYN-ROW-02", "готово"],
                ["Подготовил", "SYNTHETIC-REFERENCE-NOT-A-RECORD"],
            ],
            id="prepared-by-footer",
        ),
    ],
)
def test_k19_ambiguous_repeated_headers_and_note_footers_never_become_exact_records(
    case_id: str,
    rows: list[list[str]],
) -> None:
    """A same-width tail is not proof that every row is a data record."""

    case = {
        "id": case_id,
        "format": "xlsx",
        "filename": f"synthetic-{case_id}.xlsx",
        "rows": rows,
        "question": "Сколько позиций во всей этой таблице?",
    }
    result, projected = _project(case)
    index = result.office_structure_index or {}

    assert index["complete"] is True
    assert index["record_sets"] == []
    answer = code_owned_office_answer(
        case["question"],
        projected,
        kind_override="count_records",
    )
    assert answer is not None
    assert answer["status"] == "unknown"
    assert answer["kind"] == "unavailable"


@pytest.mark.parametrize(
    ("label", "value"),
    [
        ("Дата формирования", "2026-08-08"),
        ("Сформировано", "2026-08-08"),
        ("Проверено", "SYNTHETIC-AUDITOR"),
        ("Утверждено", "SYNTHETIC-AUDITOR"),
        ("Версия отчёта", "1.0"),
        ("Дата выгрузки", "2026-08-08"),
        ("Generated at", "2026-08-08"),
        ("Approved by", "SYNTHETIC-AUDITOR"),
    ],
)
def test_k19_terminal_report_metadata_never_becomes_an_exact_record(
    label: str,
    value: str,
) -> None:
    case = {
        "id": "terminal-report-annotation",
        "format": "xlsx",
        "filename": "synthetic-terminal-report-annotation.xlsx",
        "rows": [
            ["ID", "Статус"],
            ["SYN-ROW-01", "готово"],
            ["SYN-ROW-02", "готово"],
            [label, value],
        ],
        "question": "Сколько позиций во всей этой таблице?",
    }
    result, projected = _project(case)

    assert (result.office_structure_index or {})["record_sets"] == []
    answer = code_owned_office_answer(
        case["question"],
        projected,
        kind_override="count_records",
    )
    assert answer is not None
    assert answer["status"] == "unknown"
    assert answer["kind"] == "unavailable"
